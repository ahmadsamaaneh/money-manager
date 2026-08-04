"""Excel database layer for Money Manager."""

from __future__ import annotations

import os
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

DATA_DIR = Path(os.environ.get("DATA_DIR", Path(__file__).resolve().parent / "data"))
EXCEL_PATH = DATA_DIR / "money.xlsx"

SHEET_SETTINGS = "الإعدادات"
SHEET_TRANSACTIONS = "الحركات"
SHEET_OBLIGATIONS = "الالتزامات"
SHEET_LOANS = "الديون"

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RTL_ALIGN = Alignment(horizontal="right")


def _style_header(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(1, col)
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = RTL_ALIGN


def _ensure_workbook() -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if EXCEL_PATH.exists():
        return EXCEL_PATH

    wb = Workbook()

    ws = wb.active
    ws.title = SHEET_SETTINGS
    ws.append(["المفتاح", "القيمة"])
    _style_header(ws, 2)
    ws.append(["total_balance", 0])
    ws.append(["daily_limit", 50])
    ws.append(["spend_budget", 0])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    wt = wb.create_sheet(SHEET_TRANSACTIONS)
    wt.append(["المعرف", "التاريخ", "النوع", "المبلغ", "الاسم", "الرصيد_بعد"])
    _style_header(wt, 6)
    for col, width in enumerate([10, 20, 12, 14, 28, 14], start=1):
        wt.column_dimensions[chr(64 + col)].width = width

    wo = wb.create_sheet(SHEET_OBLIGATIONS)
    wo.append(
        [
            "المعرف",
            "الاسم",
            "المبلغ",
            "يوم_الاستحقاق",
            "ملاحظات",
            "نشط",
            "آخر_تنبيه",
            "آخر_دفع",
        ]
    )
    _style_header(wo, 8)
    for col, width in enumerate([10, 24, 14, 14, 28, 10, 18, 18], start=1):
        wo.column_dimensions[chr(64 + col)].width = width

    _create_loans_sheet(wb)
    wb.save(EXCEL_PATH)
    return EXCEL_PATH


def _ensure_obligations_schema(wb) -> None:
    ws = wb[SHEET_OBLIGATIONS]
    headers = [cell.value for cell in ws[1]]
    if "آخر_دفع" not in headers:
        col = len(headers) + 1
        cell = ws.cell(1, col, "آخر_دفع")
        cell.fill = copy(HEADER_FILL)
        cell.font = copy(HEADER_FONT)
        cell.alignment = RTL_ALIGN
        ws.column_dimensions[chr(64 + col)].width = 18


def _create_loans_sheet(wb) -> None:
    if SHEET_LOANS in wb.sheetnames:
        return
    wl = wb.create_sheet(SHEET_LOANS)
    wl.append(
        [
            "المعرف",
            "الاسم",
            "المبلغ_الأصلي",
            "المتبقي",
            "طريقة_الرجوع",
            "الفترة",
            "ملاحظات",
            "نشط",
            "تاريخ_الإنشاء",
        ]
    )
    _style_header(wl, 9)
    for col, width in enumerate([10, 22, 14, 12, 14, 16, 24, 10, 20], start=1):
        wl.column_dimensions[chr(64 + col)].width = width


def _load():
    _ensure_workbook()
    wb = load_workbook(EXCEL_PATH)
    changed = False
    if SHEET_LOANS not in wb.sheetnames:
        _create_loans_sheet(wb)
        changed = True
    if SHEET_OBLIGATIONS in wb.sheetnames:
        before = [cell.value for cell in wb[SHEET_OBLIGATIONS][1]]
        _ensure_obligations_schema(wb)
        after = [cell.value for cell in wb[SHEET_OBLIGATIONS][1]]
        if before != after:
            changed = True
    if changed:
        _save(wb)
    return wb


def _save(wb) -> None:
    wb.save(EXCEL_PATH)


def _settings_map(wb) -> dict[str, Any]:
    ws = wb[SHEET_SETTINGS]
    data: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        data[str(row[0])] = row[1]
    return data


def _set_setting(wb, key: str, value: Any) -> None:
    ws = wb[SHEET_SETTINGS]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == key:
            row[1].value = value
            return
    ws.append([key, value])


def _next_id(ws) -> int:
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            try:
                max_id = max(max_id, int(row[0]))
            except (TypeError, ValueError):
                continue
    return max_id + 1


def get_summary() -> dict[str, Any]:
    wb = _load()
    settings = _settings_map(wb)
    total = float(settings.get("total_balance") or 0)
    daily_limit = float(settings.get("daily_limit") or 0)
    spend_budget = float(settings.get("spend_budget") or 0)

    today = date.today().isoformat()
    spent_today = 0.0
    transactions: list[dict[str, Any]] = []

    ws = wb[SHEET_TRANSACTIONS]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        tx_date = str(row[1])[:10] if row[1] else ""
        tx_type = str(row[2] or "")
        amount = float(row[3] or 0)
        item = {
            "id": int(row[0]),
            "date": str(row[1] or ""),
            "type": tx_type,
            "amount": amount,
            "name": str(row[4] or ""),
            "balance_after": float(row[5] or 0),
        }
        transactions.append(item)
        if tx_type == "مصروف" and tx_date == today:
            spent_today += amount

    transactions.reverse()
    remaining_today = max(daily_limit - spent_today, 0)
    total_lent = sum(item["remaining"] for item in list_loans(wb))
    return {
        "total_balance": total,
        "daily_limit": daily_limit,
        "spend_budget": spend_budget,
        "spent_today": spent_today,
        "remaining_today": remaining_today,
        "total_lent": total_lent,
        "transactions": transactions[:100],
        "excel_path": str(EXCEL_PATH),
    }


def deposit(amount: float, name: str = "إيداع") -> dict[str, Any]:
    if amount <= 0:
        raise ValueError("المبلغ يجب أن يكون أكبر من صفر")

    wb = _load()
    settings = _settings_map(wb)
    total = float(settings.get("total_balance") or 0) + amount
    _set_setting(wb, "total_balance", total)

    ws = wb[SHEET_TRANSACTIONS]
    tx_id = _next_id(ws)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([tx_id, now, "إيداع", amount, name.strip() or "إيداع", total])
    _save(wb)
    return get_summary()


def set_total_balance(amount: float, note: str = "تصحيح رصيد") -> dict[str, Any]:
    if amount < 0:
        raise ValueError("رأس المال لا يمكن أن يكون سالباً")

    wb = _load()
    settings = _settings_map(wb)
    old_total = float(settings.get("total_balance") or 0)
    spend_budget = float(settings.get("spend_budget") or 0)

    _set_setting(wb, "total_balance", amount)
    if spend_budget > amount:
        _set_setting(wb, "spend_budget", amount)

    ws = wb[SHEET_TRANSACTIONS]
    tx_id = _next_id(ws)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    diff = amount - old_total
    label = (note.strip() or "تصحيح رصيد") + f" (من {old_total:.2f})"
    ws.append([tx_id, now, "تعديل", abs(diff), label, amount])
    _save(wb)
    return get_summary()


def set_spend_budget(amount: float) -> dict[str, Any]:
    if amount < 0:
        raise ValueError("المبلغ لا يمكن أن يكون سالباً")

    wb = _load()
    settings = _settings_map(wb)
    total = float(settings.get("total_balance") or 0)
    if amount > total:
        raise ValueError("مبلغ الصرف أكبر من إجمالي المال")

    _set_setting(wb, "spend_budget", amount)
    _save(wb)
    return get_summary()


def set_daily_limit(amount: float) -> dict[str, Any]:
    if amount < 0:
        raise ValueError("السقف اليومي لا يمكن أن يكون سالباً")

    wb = _load()
    _set_setting(wb, "daily_limit", amount)
    _save(wb)
    return get_summary()


def add_expense(amount: float, name: str) -> dict[str, Any]:
    if amount <= 0:
        raise ValueError("المبلغ يجب أن يكون أكبر من صفر")
    if not name.strip():
        raise ValueError("اكتب اسم المصروف أو سبب الشراء")

    wb = _load()
    settings = _settings_map(wb)
    total = float(settings.get("total_balance") or 0)
    daily_limit = float(settings.get("daily_limit") or 0)
    spend_budget = float(settings.get("spend_budget") or 0)

    if amount > total:
        raise ValueError("الرصيد غير كافٍ")

    today = date.today().isoformat()
    spent_today = 0.0
    ws = wb[SHEET_TRANSACTIONS]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        tx_date = str(row[1])[:10] if row[1] else ""
        if str(row[2] or "") == "مصروف" and tx_date == today:
            spent_today += float(row[3] or 0)

    if spent_today + amount > daily_limit:
        remaining = max(daily_limit - spent_today, 0)
        raise ValueError(f"تجاوزت السقف اليومي. المتبقي اليوم: {remaining:.2f}")

    if spend_budget > 0 and amount > spend_budget:
        raise ValueError("المبلغ أكبر من مبلغ الصرف المخصص")

    total -= amount
    if spend_budget > 0:
        spend_budget = max(spend_budget - amount, 0)
        _set_setting(wb, "spend_budget", spend_budget)

    _set_setting(wb, "total_balance", total)
    tx_id = _next_id(ws)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([tx_id, now, "مصروف", amount, name.strip(), total])
    _save(wb)
    return get_summary()


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()[:10]
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _add_month(d: date, day: int) -> date:
    day = max(1, min(int(day), 28))
    if d.month == 12:
        return date(d.year + 1, 1, day)
    return date(d.year, d.month + 1, day)


def _due_date_for_day(day: int, today: date | None = None, last_paid: date | None = None) -> date:
    today = today or date.today()
    day = max(1, min(int(day), 28))

    if last_paid is not None:
        return _add_month(last_paid, day)

    this_month = date(today.year, today.month, day)
    if this_month >= today:
        return this_month
    return _add_month(today, day)


def list_obligations() -> list[dict[str, Any]]:
    wb = _load()
    ws = wb[SHEET_OBLIGATIONS]
    today = date.today()
    items: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        active = str(row[5] if row[5] is not None else "1").strip()
        if active in {"0", "False", "false", "لا"}:
            continue

        due_day = int(row[3] or 1)
        last_paid = _parse_date(row[7] if len(row) > 7 else None)
        due = _due_date_for_day(due_day, today, last_paid)
        days_left = (due - today).days
        items.append(
            {
                "id": int(row[0]),
                "name": str(row[1] or ""),
                "amount": float(row[2] or 0),
                "due_day": due_day,
                "notes": str(row[4] or ""),
                "due_date": due.isoformat(),
                "days_left": days_left,
                "urgent": days_left <= 3,
                "last_paid": last_paid.isoformat() if last_paid else "",
            }
        )

    items.sort(key=lambda x: x["days_left"])
    return items


def add_obligation(name: str, amount: float, due_day: int, notes: str = "") -> list[dict[str, Any]]:
    if not name.strip():
        raise ValueError("اكتب اسم الالتزام")
    if amount <= 0:
        raise ValueError("المبلغ يجب أن يكون أكبر من صفر")
    day = int(due_day)
    if day < 1 or day > 28:
        raise ValueError("يوم الاستحقاق بين 1 و 28")

    wb = _load()
    ws = wb[SHEET_OBLIGATIONS]
    oid = _next_id(ws)
    ws.append([oid, name.strip(), amount, day, notes.strip(), 1, "", ""])
    _save(wb)
    return list_obligations()


def delete_obligation(obligation_id: int) -> list[dict[str, Any]]:
    wb = _load()
    ws = wb[SHEET_OBLIGATIONS]
    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None and int(row[0].value) == int(obligation_id):
            row[5].value = 0
            _save(wb)
            return list_obligations()
    raise ValueError("الالتزام غير موجود")


def pay_obligation(obligation_id: int) -> dict[str, Any]:
    wb = _load()
    ws = wb[SHEET_OBLIGATIONS]
    today = date.today()
    target = None

    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None and int(row[0].value) == int(obligation_id):
            target = row
            break

    if target is None:
        raise ValueError("الالتزام غير موجود")

    active = str(target[5].value if target[5].value is not None else "1").strip()
    if active in {"0", "False", "false", "لا"}:
        raise ValueError("الالتزام غير نشط")

    name = str(target[1].value or "التزام")
    amount = float(target[2].value or 0)
    due_day = int(target[3].value or 1)
    last_paid = _parse_date(target[7].value if len(target) > 7 else None)
    due = _due_date_for_day(due_day, today, last_paid)

    if amount <= 0:
        raise ValueError("مبلغ الالتزام غير صالح")

    settings = _settings_map(wb)
    total = float(settings.get("total_balance") or 0)
    spend_budget = float(settings.get("spend_budget") or 0)

    if amount > total:
        raise ValueError("الرصيد غير كافٍ لدفع هذا الالتزام")

    total -= amount
    if spend_budget > total:
        _set_setting(wb, "spend_budget", total)
    _set_setting(wb, "total_balance", total)

    # Mark this cycle as paid so next due moves to the following month.
    target[7].value = due.isoformat()

    tx = wb[SHEET_TRANSACTIONS]
    tx_id = _next_id(tx)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tx.append([tx_id, now, "التزام", amount, f"دفع «{name}»", total])
    _save(wb)
    return {
        "summary": get_summary(),
        "obligations": list_obligations(),
        "notifications": get_notifications(7),
    }


def get_notifications(within_days: int = 7) -> list[dict[str, Any]]:
    notes = []
    for item in list_obligations():
        if item["days_left"] <= within_days:
            if item["days_left"] == 0:
                message = f"اليوم استحقاق «{item['name']}» بمبلغ {item['amount']:.2f}"
            elif item["days_left"] == 1:
                message = f"غداً استحقاق «{item['name']}» بمبلغ {item['amount']:.2f}"
            else:
                message = (
                    f"باقي {item['days_left']} أيام على «{item['name']}» "
                    f"({item['amount']:.2f})"
                )
            notes.append(
                {
                    "id": item["id"],
                    "title": item["name"],
                    "message": message,
                    "days_left": item["days_left"],
                    "amount": item["amount"],
                    "due_date": item["due_date"],
                    "urgent": item["urgent"],
                }
            )
    return notes


def emergency_withdraw(amount: float, name: str = "سحب طوارئ") -> dict[str, Any]:
    if amount <= 0:
        raise ValueError("المبلغ يجب أن يكون أكبر من صفر")

    wb = _load()
    settings = _settings_map(wb)
    total = float(settings.get("total_balance") or 0)
    spend_budget = float(settings.get("spend_budget") or 0)

    if amount > total:
        raise ValueError("الرصيد غير كافٍ")

    total -= amount
    if spend_budget > total:
        _set_setting(wb, "spend_budget", total)

    _set_setting(wb, "total_balance", total)
    ws = wb[SHEET_TRANSACTIONS]
    tx_id = _next_id(ws)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    label = name.strip() or "سحب طوارئ"
    ws.append([tx_id, now, "طوارئ", amount, label, total])
    _save(wb)
    return get_summary()


def list_loans(wb=None) -> list[dict[str, Any]]:
    own_wb = wb is None
    if own_wb:
        wb = _load()
    ws = wb[SHEET_LOANS]
    items: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        active = str(row[7] if len(row) > 7 and row[7] is not None else "1").strip()
        if active in {"0", "False", "false", "لا"}:
            continue

        mode = str(row[4] or "مرة واحدة")
        items.append(
            {
                "id": int(row[0]),
                "name": str(row[1] or ""),
                "amount": float(row[2] or 0),
                "remaining": float(row[3] or 0),
                "mode": mode,
                "period": str(row[5] or ""),
                "notes": str(row[6] or ""),
                "created_at": str(row[8] or "") if len(row) > 8 else "",
            }
        )

    items.sort(key=lambda x: x["remaining"], reverse=True)
    return items


def add_loan(
    name: str,
    amount: float,
    mode: str = "مرة واحدة",
    period: str = "",
    notes: str = "",
) -> dict[str, Any]:
    if not name.strip():
        raise ValueError("اكتب اسم الشخص")
    if amount <= 0:
        raise ValueError("المبلغ يجب أن يكون أكبر من صفر")

    mode = mode.strip()
    if mode not in {"مرة واحدة", "على فترات"}:
        raise ValueError("طريقة الرجوع غير صحيحة")
    if mode == "على فترات" and not str(period).strip():
        raise ValueError("اكتب الفترة التقريبية مثل: كل أسبوع أو كل شهر")

    wb = _load()
    settings = _settings_map(wb)
    total = float(settings.get("total_balance") or 0)
    spend_budget = float(settings.get("spend_budget") or 0)

    if amount > total:
        raise ValueError("الرصيد غير كافٍ لإقراض هذا المبلغ")

    total -= amount
    if spend_budget > total:
        _set_setting(wb, "spend_budget", total)
    _set_setting(wb, "total_balance", total)

    ws = wb[SHEET_LOANS]
    loan_id = _next_id(ws)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append(
        [
            loan_id,
            name.strip(),
            amount,
            amount,
            mode,
            period.strip(),
            notes.strip(),
            1,
            now,
        ]
    )

    tx = wb[SHEET_TRANSACTIONS]
    tx_id = _next_id(tx)
    label = f"سلفة لـ {name.strip()}"
    tx.append([tx_id, now, "سلفة", amount, label, total])
    _save(wb)
    return {"summary": get_summary(), "loans": list_loans()}


def repay_loan(loan_id: int, amount: float) -> dict[str, Any]:
    if amount <= 0:
        raise ValueError("المبلغ يجب أن يكون أكبر من صفر")

    wb = _load()
    ws = wb[SHEET_LOANS]
    target = None
    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None and int(row[0].value) == int(loan_id):
            target = row
            break

    if target is None:
        raise ValueError("السلفة غير موجودة")

    active = str(target[7].value if target[7].value is not None else "1").strip()
    if active in {"0", "False", "false", "لا"}:
        raise ValueError("هذه السلفة مغلقة")

    remaining = float(target[3].value or 0)
    if amount > remaining:
        raise ValueError(f"المبلغ أكبر من المتبقي ({remaining:.2f})")

    person = str(target[1].value or "شخص")
    remaining -= amount
    target[3].value = remaining
    if remaining <= 0:
        target[7].value = 0

    settings = _settings_map(wb)
    total = float(settings.get("total_balance") or 0) + amount
    _set_setting(wb, "total_balance", total)

    tx = wb[SHEET_TRANSACTIONS]
    tx_id = _next_id(tx)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    label = f"تسديد من {person}"
    if remaining <= 0:
        label += " (اكتملت)"
    tx.append([tx_id, now, "تسديد", amount, label, total])
    _save(wb)
    return {"summary": get_summary(), "loans": list_loans()}


def delete_loan(loan_id: int) -> list[dict[str, Any]]:
    wb = _load()
    ws = wb[SHEET_LOANS]
    for row in ws.iter_rows(min_row=2):
        if row[0].value is not None and int(row[0].value) == int(loan_id):
            row[7].value = 0
            _save(wb)
            return list_loans()
    raise ValueError("السلفة غير موجودة")
